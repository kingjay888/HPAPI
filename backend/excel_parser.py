from io import BytesIO
from typing import Any

from openpyxl import load_workbook

PRODUCT_COLUMN_ALIASES = {
    "product",
    "product number",
    "product_number",
    "productnumber",
    "sku",
    "part number",
    "part_number",
    "partnumber",
    "model",
    "device",
    "printer",
    "hp product",
    "hp product number",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _normalize_product(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def parse_product_list(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_normalize_header(cell) for cell in rows[0]]
    product_col = None

    for index, name in enumerate(header):
        if name in PRODUCT_COLUMN_ALIASES:
            product_col = index
            break

    # If no recognized header, treat first column as product numbers.
    if product_col is None:
        product_col = 0
        data_rows = rows
    else:
        data_rows = rows[1:]

    products: list[str] = []
    seen: set[str] = set()

    for row in data_rows:
        if not row or product_col >= len(row):
            continue
        product = _normalize_product(row[product_col])
        if not product:
            continue
        key = product.upper()
        if key in seen:
            continue
        seen.add(key)
        products.append(product)

    return products
